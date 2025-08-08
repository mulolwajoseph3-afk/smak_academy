from django.shortcuts import render, redirect, HttpResponseRedirect
from django.core.exceptions import ObjectDoesNotExist
from django.db import DatabaseError
from django.shortcuts import get_object_or_404
from django.contrib import messages
from .forms import StudentRegistration,AutreFraisForm, ClassesRegistratation,TauxRegistration,AnneeScolaireRegistration, PaiementRegistration, RegisterForm, UserLoginForm, MoisClasseFilterForm, ClasseFilterForm
from .models import student, Classes, AnneeScolaire, Paiement, AutreFrais, Depense, Taux
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import openpyxl
import barcode
from io import BytesIO
from django.http import HttpResponse
from barcode.writer import ImageWriter
from django.core.files.storage import default_storage
import base64
from barcode import Code128
from openpyxl.styles import Font
from django.http import HttpResponse
from decimal import Decimal
from django.utils import timezone
from django.utils.timezone import now
from datetime import datetime
from datetime import date
from django.db.models import Sum
import xlwt
import logging
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
logger = logging.getLogger(__name__)

# Create your views here.

#creer compte  
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden

# LES CONNEXION POUR PROMO , COMPTABLE , SUPER ADMIN

@login_required
def register(request):
    if request.user.role != 'promo':
        return HttpResponseForbidden("Vous n'êtes pas autorisé à créer un compte comptable.")

    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)

            # 🔐 Hash du mot de passe
            user.set_password(form.cleaned_data['password'])

            # 🎯 Rôle et rattachement automatique
            user.role = 'comptable'
            user.ecole = request.user.ecole

            # 🧼 Gestion de l'email vide
            email = form.cleaned_data.get('email')
            user.email = email.strip() if email and email.strip() != '' else None

            user.save()
            messages.success(request, "Le compte comptable a bien été créé.")
            return redirect('dashboard_promo')  # Adapte le nom si besoin
    else:
        form = RegisterForm()

    return render(request, 'promo/register.html', {'form': form})


# Vue de connexion
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.shortcuts import render, redirect
from .forms import UserLoginForm

from academy.models import AnneeScolaire


def login_view(request):
    if request.method == 'POST':
        form = UserLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)

            # ✅ Initialiser année scolaire pour tous les rôles liés à une école
            if user.ecole:
                annees = AnneeScolaire.objects.filter(ecole=user.ecole, active=True)

                if annees.count() == 1:
                    request.session['annee_scolaire'] = annees.first().id
                elif annees.exists():
                    # 🔀 Plusieurs années actives → redirection pour choisir
                    return redirect('changer_annee')
                else:
                    messages.error(request, "Aucune année scolaire active n’a été définie pour votre école.")
                    return redirect('changer_annee')

            # 🔁 Redirection selon le rôle
            if user.role == 'superadmin':
                return redirect('superadmin_dashboard')
            else:
                return redirect('index_promo' if user.role == 'promo' else 'index')

        else:
            messages.error(request, "Formulaire invalide")
    else:
        form = UserLoginForm()

    return render(request, 'login.html', {'form': form})

#se deconnecter 
def logout_view(request):
    logout(request)  # Déconnecte l'utilisateur
    return redirect('login')  # Redirige vers la page de login après la déconnexion

# VUE BASHBOARD POUR COMPTABLE
@login_required
def index(request):
    if request.user.role != 'comptable':
        return redirect('login')

    annee_id = request.session.get('annee_scolaire')
    annee_selectionnee_obj = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole
    ).first()

    if not annee_selectionnee_obj:
        annee_selectionnee_obj = AnneeScolaire.objects.filter(
            ecole=request.user.ecole, active=True
        ).first()
        if annee_selectionnee_obj:
            request.session['annee_scolaire'] = annee_selectionnee_obj.id
            messages.info(request, f"Sélection automatique de l’année : {annee_selectionnee_obj.nom_annee}")
        else:
            messages.error(request, "Aucune année scolaire active disponible.")
            return redirect('logout')

    name_annee_selectionnee = annee_selectionnee_obj.nom_annee

    # 👨🎓 Statistiques
    number_student = student.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        created_by=request.user
    ).count()

    number_man = student.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        sexe='M',
        created_by=request.user
    ).count()

    number_girl = student.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        sexe='F',
        created_by=request.user
    ).count()

    number_class = Classes.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        created_by=request.user
    ).count()

    # 💰 Totaux financiers : Encaissements réels
    total_revenu_cdf = Paiement.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        recu=True,
        created_by=request.user
    ).aggregate(total=Sum('montant_cdf_brut'))['total'] or Decimal('0')

    total_revenu_usd = Paiement.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        recu=True,
        created_by=request.user
    ).aggregate(total=Sum('montant_usd_brut'))['total'] or Decimal('0')

    total_depense_cdf = Depense.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        created_by=request.user
    ).aggregate(total=Sum('montant'))['total'] or Decimal('0')

    total_depense_usd = Depense.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        created_by=request.user
    ).aggregate(total=Sum('montant_usd'))['total'] or Decimal('0')

    # 📅 Mouvements du jour
    aujourd_hui = timezone.now().date()

    revenu_cdf_jour = Paiement.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        date_paiement=aujourd_hui,
        recu=True,
        created_by=request.user
    ).aggregate(total=Sum('montant_cdf_brut'))['total'] or Decimal('0')

    revenu_usd_jour = Paiement.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        date_paiement=aujourd_hui,
        recu=True,
        created_by=request.user
    ).aggregate(total=Sum('montant_usd_brut'))['total'] or Decimal('0')

    depense_cdf_jour = Depense.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        date_depense=aujourd_hui,
        created_by=request.user
    ).aggregate(total=Sum('montant'))['total'] or Decimal('0')

    depense_usd_jour = Depense.objects.filter(
        annee_scolaire=annee_selectionnee_obj,
        date_depense=aujourd_hui,
        created_by=request.user
    ).aggregate(total=Sum('montant_usd'))['total'] or Decimal('0')

    annees = AnneeScolaire.objects.filter(
        ecole=request.user.ecole,
        active=True
    ).order_by('-nom_annee')

    context = {
        'number_student': number_student,
        'number_man': number_man,
        'number_girl': number_girl,
        'number_class': number_class,
        'annee_selectionnee': annee_selectionnee_obj,
        'name_annee_selectionnee': name_annee_selectionnee,
        'user': request.user,
        'total_revenu_cdf': total_revenu_cdf,
        'total_revenu_usd': total_revenu_usd,
        'total_depense_cdf': total_depense_cdf,
        'total_depense_usd': total_depense_usd,
        'revenu_cdf_jour': revenu_cdf_jour,
        'revenu_usd_jour': revenu_usd_jour,
        'depense_cdf_jour': depense_cdf_jour,
        'depense_usd_jour': depense_usd_jour,
        'annees': annees
    }

    return render(request, 'index.html', context)

#VUE CHANGER ANNEE POUR PROMO ET COMPTABLE PAR ECOLE 
@login_required
def changer_annee(request):
    if request.user.role not in ['promo', 'comptable']:
        return HttpResponseForbidden("Accès refusé")

    annees_disponibles = AnneeScolaire.objects.filter(
        ecole=request.user.ecole,
        active=True
    ).order_by('nom_annee')

    # 🎯 Si le formulaire est soumis
    if request.method == 'POST':
        annee_id = request.POST.get('annee_scolaire')
        try:
            annee_scolaire = annees_disponibles.get(id=annee_id)
            request.session['annee_scolaire'] = annee_scolaire.id
            messages.success(request, f"Année scolaire « {annee_scolaire.nom_annee} » sélectionnée.")
        except AnneeScolaire.DoesNotExist:
            messages.error(request, "Année scolaire invalide.")

        # ✅ Redirection directe
        return redirect('index_promo' if request.user.role == 'promo' else 'index')

    # 🔁 Si on arrive en GET et qu’il y a au moins une année
    elif annees_disponibles.exists():
        annee_scolaire = annees_disponibles.first()
        request.session['annee_scolaire'] = annee_scolaire.id
        messages.info(request, f"Année scolaire « {annee_scolaire.nom_annee} » sélectionnée par défaut.")
        return redirect('index_promo' if request.user.role == 'promo' else 'index')

    # ❌ Aucun année dispo
    messages.error(request, "Aucune année scolaire active disponible. Contactez le superadmin.")
    return redirect('login')

#VUE INSCRIPTION POUR COMPTABLE 

@login_required
def get_inscription_fee(request):
    if request.user.role != 'comptable':
        return JsonResponse({})  # Accès interdit silencieusement

    devise = request.GET.get('devise')
    annee_id = request.session.get('annee_scolaire')

    frais = AutreFrais.objects.filter(
        description='inscription',
        annee_scolaire_id=annee_id,
        created_by=request.user  # 🔐 Séparation par comptable
    ).first()

    if frais:
        if devise == 'cdf':
            return JsonResponse({'montant': float(frais.montant_cdf)})
        elif devise == 'usd':
            return JsonResponse({'montant': float(frais.montant_usd)})

    return JsonResponse({})


@login_required
def inscription(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès non autorisé")

    annee_id = request.session.get('annee_scolaire')
    annee_active = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True  # 🔐 Validation par activation
    ).first()

    if not annee_active:
        messages.warning(request, "Veuillez d'abord sélectionner une année scolaire valide.")
        return redirect('index')

    # 👦 Élèves créés par ce comptable uniquement
    eleves = student.objects.filter(
        annee_scolaire=annee_active,
        created_by=request.user
    )
    number_student = eleves.count()

    if request.method == 'POST':
        fm = StudentRegistration(request.POST or None, annee_scolaire_id=annee_active.id, created_by=request.user)
        if fm.is_valid():
            eleve = fm.save(commit=False)
            eleve.created_by = request.user  # 🔐 Tag comptable
            eleve.annee_scolaire = annee_active
            eleve.save()

            eleve = fm.save()
            paiement = fm.paiement_created

            if paiement:
                messages.success(request, "Élève inscrit et paiement d'inscription enregistré.")
                return redirect('receipt_view', paiement_id=paiement.id)
            else:
                messages.warning(request, "Élève enregistré, mais aucun paiement généré.")
                return redirect('index')
    else:
        fm = StudentRegistration(annee_scolaire_id=annee_active.id)

    context = {
        'eleves': eleves,
        'form': fm,
        'number_student': number_student
    }
    return render(request, 'inscription.html', context)

@login_required
def UpdateStudent(request, id):
    st = get_object_or_404(student, pk=id)

    # 🔒 Vérifie que le comptable est bien le créateur de cet élève
    if request.user.role != 'comptable' or st.created_by != request.user:
        return HttpResponseForbidden("Modification non autorisée")

    if request.method == 'POST':
        fm = StudentRegistration(request.POST, instance=st)
        if fm.is_valid():
            updated_student = fm.save(commit=False)
            updated_student.created_by = request.user  # Sécurité conservée
            updated_student.save()
            messages.success(request, "Élève mis à jour avec succès.")
            return redirect('details_classe')
        else:
            messages.error(request, "Erreur dans le formulaire. Veuillez corriger les champs.")
    else:
        fm = StudentRegistration(instance=st)

    return render(request, 'updatestudent.html', {'form': fm, 'st': st})

@login_required
def DeleteStudent(request, id):
    st = get_object_or_404(student, pk=id)

    # 🔒 Sécurité : seul le comptable qui a inscrit l’élève peut le supprimer
    if request.user.role != 'comptable' or st.created_by != request.user:
        return HttpResponseForbidden("Suppression non autorisée")

    if request.method == 'POST':
        st.delete()
        messages.warning(request, "Élève supprimé.")
        return redirect('details_classe')

    return render(request, 'confirm_delete_student.html', {'eleve': st})

@login_required
def Details_Student(request, id):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    try:
        st = student.objects.get(pk=id)

        # 🔐 Vérification : le comptable peut uniquement voir ses propres élèves
        if st.created_by != request.user:
            return HttpResponseForbidden("Cet élève ne vous appartient pas")

        # 🔍 Paiements enregistrés par ce comptable pour cet élève
        paiements = Paiement.objects.filter(
            eleve_id=id,
            annee_scolaire=st.annee_scolaire,
            created_by=request.user
        )

        content = {
            'st': st,
            'paiements': paiements
        }
        return render(request, 'detail_student.html', content)

    except student.DoesNotExist:
        return render(request, 'error.html', {'message': "L'élève n'existe pas"})


#TAUX

@login_required
def enregistrer_taux(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès non autorisé")

    annee_id = request.session.get('annee_scolaire')
    annee = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee:
        messages.error(request, "Aucune année scolaire valide sélectionnée.")
        return redirect('index')

    form = TauxRegistration(request.POST or None)

    if request.method == 'POST' and form.is_valid():
        taux_instance = form.save(commit=False)
        taux_instance.annee_scolaire = annee
        taux_instance.created_by = request.user  # 🔒 Attribution comptable
        taux_instance.save()
        messages.success(request, "Taux enregistré avec succès.")
        return redirect('ajout_taux')

    taux_list = Taux.objects.filter(
        annee_scolaire=annee,
        created_by=request.user  # 🔍 Taux du comptable courant uniquement
    ).order_by('-date_taux')

    return render(request, 'taux_registration.html', {
        'form': form,
        'taux_list': taux_list,
        'annee': annee
    })
    
@login_required
def modifier_taux(request, taux_id):
    taux_instance = get_object_or_404(Taux, id=taux_id)

    # 🔐 Sécurité : le taux doit appartenir au comptable connecté
    if taux_instance.created_by != request.user:
        return HttpResponseForbidden("Modification non autorisée")

    form = TauxRegistration(request.POST or None, instance=taux_instance)

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Taux modifié.")
        return redirect('ajout_taux')

    annee = taux_instance.annee_scolaire
    taux_list = Taux.objects.filter(
        annee_scolaire=annee,
        created_by=request.user
    ).order_by('-date_taux')

    return render(request, 'taux_registration.html', {
        'form': form,
        'taux_list': taux_list,
        'annee': annee,
        'taux_instance': taux_instance
    })
    
@login_required
def supprimer_taux(request, taux_id):
    taux_instance = get_object_or_404(Taux, id=taux_id)

    if taux_instance.created_by != request.user:
        return HttpResponseForbidden("Suppression non autorisée")

    if request.method == 'POST':
        taux_instance.delete()
        messages.warning(request, "Taux supprimé.")
        return redirect('ajout_taux')

    return render(request, 'confirmation_suppression_taux.html', {'taux': taux_instance})

#VUE POUR PAIEMENT FRAIS COMPTABLE
from django.shortcuts import render, redirect
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.contrib import messages
from decimal import Decimal
from .forms import PaiementRegistration
from .models import Paiement, student, Classes, AnneeScolaire, AutreFrais, Taux

@login_required
def paiements(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès non autorisé.")

    annee_scolaire_id = request.session.get('annee_scolaire')
    annee_active = AnneeScolaire.objects.filter(
        id=annee_scolaire_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_active:
        messages.warning(request, "Veuillez d'abord sélectionner une année scolaire valide.")
        return redirect('index')

    paiements = Paiement.objects.filter(
        annee_scolaire=annee_active,
        created_by=request.user
    )

    fm = PaiementRegistration(
        request.POST or None,
        annee_scolaire_id=annee_active.id,
        created_by=request.user
    )

    if request.method == 'POST' and fm.is_valid():
        paiement = fm.save(commit=False)
        eleve = fm.cleaned_data['eleve']
        type_paiement = fm.cleaned_data['type_paiement']
        mois = fm.cleaned_data['mois']
        devise = fm.cleaned_data.get('devise')
        montant = fm.cleaned_data.get('montant')
        avance = fm.cleaned_data.get('avance')
        classe = eleve.classe
        paiement.devise = devise

        montant_minerval = classe.montant if devise == 'cdf' else classe.montant_usd
        taux_instance = Taux.objects.filter(
            annee_scolaire=annee_active,
            created_by=request.user
        ).order_by('-date_taux').first()
        taux_valeur = taux_instance.valeur if taux_instance else Decimal('2500')

        montant_usd = Decimal(request.POST.get('montant_usd') or '0')
        montant_cdf = Decimal(request.POST.get('montant_cdf') or '0')
        avance_brute = Decimal(avance) if avance not in [None, ''] else Decimal(montant or '0')

        montant_usd_reparti = Decimal('0')
        montant_cdf_reparti = Decimal('0')

        # 💵 Encaissement initial selon la devise
        if devise == 'mixte':
            AnneeScolaire.objects.filter(id=annee_active.id).update(
                total_revenu=F('total_revenu') + montant_cdf,
                total_revenu_usd=F('total_revenu_usd') + montant_usd
            )
            annee_active.refresh_from_db()

            avance_mixte = montant_cdf + (montant_usd * taux_valeur)
            paiement.avance = avance_mixte
            paiement.montant_usd_brut = montant_usd
            paiement.montant_cdf_brut = montant_cdf
            paiement.taux = taux_instance
            paiement.solde = max(montant_minerval - avance_mixte, Decimal('0'))

        elif devise == 'usd':
            paiement.avance = avance_brute * taux_valeur
            paiement.montant_usd_brut = avance_brute
            paiement.montant_cdf_brut = Decimal('0')
            paiement.taux = taux_instance
            paiement.solde = max(montant_minerval - paiement.avance, Decimal('0'))

        elif devise == 'cdf':
            paiement.avance = avance_brute
            paiement.montant_cdf_brut = avance_brute
            paiement.montant_usd_brut = Decimal('0')
            paiement.taux = None
            paiement.solde = max(montant_minerval - avance_brute, Decimal('0'))

        paiement.montant = montant_minerval

        # 🔐 Mise à jour caisse globale (une seule fois)
        if paiement.avance > 0:
            if devise == 'usd':
                AnneeScolaire.objects.filter(id=annee_active.id).update(
                    total_revenu_usd=F('total_revenu_usd') + paiement.montant_usd_brut
                )
            elif devise == 'cdf':
                AnneeScolaire.objects.filter(id=annee_active.id).update(
                    total_revenu=F('total_revenu') + paiement.montant_cdf_brut
                )
            elif devise == 'mixte':
                AnneeScolaire.objects.filter(id=annee_active.id).update(
                    total_revenu=F('total_revenu') + paiement.montant_cdf_brut,
                    total_revenu_usd=F('total_revenu_usd') + paiement.montant_usd_brut
                )
        if classe.annee_scolaire != annee_active or classe.annee_scolaire.ecole != request.user.ecole:
            fm.add_error('eleve', "L'élève ne correspond pas à l'année scolaire sélectionnée.")
        else:
            mois_ordre = ['septembre', 'octobre', 'novembre', 'décembre',
                          'janvier', 'février', 'mars', 'avril', 'mai', 'juin']

            if type_paiement == 'autre_frais':
                paiement.object_paiement = fm.cleaned_data.get('object_paiement')
                paiement.recu = True
                paiement.date_paiement = timezone.now().date()
                paiement.created_by = request.user
                paiement.classe = classe
                paiement.eleve = eleve
                paiement.annee_scolaire = annee_active
                paiement.mois = mois
                paiement.montant = montant
                paiement.solde = Decimal('0')
                paiement.save()
                return redirect('receipt_view', paiement_id=paiement.id)

            elif type_paiement == 'minerval' and mois in mois_ordre:
                index_mois = mois_ordre.index(mois)

                # ✅ Nouvelle vérification : mois déjà soldé
                paiement_deja_paye = Paiement.objects.filter(
                    eleve=eleve,
                    mois__iexact=mois,
                    type_paiement='minerval',
                    annee_scolaire=annee_active,
                    solde=Decimal('0')
                ).first()

                if paiement_deja_paye:
                    fm.add_error('mois', f"Le mois de {mois.title()} a déjà été complètement payé.")
                else:
                    for i in range(index_mois):
                        mois_precedent = mois_ordre[i]
                        paiement_solde = Paiement.objects.filter(
                            eleve=eleve,
                            type_paiement='minerval',
                            mois=mois_precedent,
                            annee_scolaire=annee_active,
                            solde=Decimal('0')
                        ).exists()
                        if not paiement_solde:
                            fm.add_error('mois', f"Le paiement du mois de {mois_precedent.title()} n’est pas encore réglé.")
                            break

                    if not fm.errors:
                        avance_restante = paiement.avance
                        paiements_crees = []

                        for i in range(index_mois, len(mois_ordre)):
                            mois_courant = mois_ordre[i]
                            if avance_restante <= 0:
                                break

                            montant_mois = classe.montant
                            paiement_mois = Paiement.objects.filter(
                                eleve=eleve,
                                mois__iexact=mois_courant,
                                type_paiement='minerval',
                                annee_scolaire=annee_active
                            ).first()

                            ajout = min(avance_restante, montant_mois if not paiement_mois else montant_mois - paiement_mois.avance)

                            if paiement_mois:
                                paiement_mois.avance += ajout
                                paiement_mois.solde = max(montant_mois - paiement_mois.avance, Decimal('0'))
                                paiement_mois.date_paiement = timezone.now().date()

                                # ✅ Répartition selon la devise utilisée POUR solder la dette, pas celle d'origine
                                if devise == 'cdf':
                                    paiement_mois.montant_cdf_brut += ajout
                                    AnneeScolaire.objects.filter(id=annee_active.id).update(
                                        total_revenu=F('total_revenu') + ajout
                                    )
                                elif devise == 'usd':
                                    paiement_mois.montant_usd_brut += ajout / taux_valeur
                                    AnneeScolaire.objects.filter(id=annee_active.id).update(
                                        total_revenu_usd=F('total_revenu_usd') + (ajout / taux_valeur)
                                    )
                                elif devise == 'mixte':
                                    part_usd = paiement.montant_usd_brut * taux_valeur
                                    part_cdf = paiement.montant_cdf_brut
                                    total_mixte = part_usd + part_cdf
                                    ratio_usd = part_usd / total_mixte if total_mixte > 0 else Decimal('0')
                                    ratio_cdf = part_cdf / total_mixte if total_mixte > 0 else Decimal('0')

                                    paiement_mois.montant_usd_brut += (ajout * ratio_usd) / taux_valeur
                                    paiement_mois.montant_cdf_brut += ajout * ratio_cdf

                                    AnneeScolaire.objects.filter(id=annee_active.id).update(
                                        total_revenu_usd=F('total_revenu_usd') + ((ajout * ratio_usd) / taux_valeur),
                                        total_revenu=F('total_revenu') + (ajout * ratio_cdf)
                                    )

                                paiement_mois.save(update_fields=[
                                    'avance', 'solde', 'date_paiement',
                                    'montant_cdf_brut', 'montant_usd_brut'
                                ])
                                avance_restante -= ajout
                                paiements_crees.append(paiement_mois)

                            elif ajout > 0:
                                montant_usd_reparti = Decimal('0')
                                montant_cdf_reparti = Decimal('0')

                                if devise == 'mixte':
                                    part_usd = paiement.montant_usd_brut * taux_valeur
                                    part_cdf = paiement.montant_cdf_brut
                                    total_mixte = part_usd + part_cdf
                                    ratio_usd = part_usd / total_mixte if total_mixte > 0 else Decimal('0')
                                    ratio_cdf = part_cdf / total_mixte if total_mixte > 0 else Decimal('0')
                                    montant_usd_reparti = (ajout * ratio_usd) / taux_valeur
                                    montant_cdf_reparti = ajout * ratio_cdf
                                elif devise == 'usd':
                                    montant_usd_reparti = ajout / taux_valeur
                                elif devise == 'cdf':
                                    montant_cdf_reparti = ajout

                                nouveau_paiement = Paiement.objects.create(
                                    eleve=eleve,
                                    classe=classe,
                                    annee_scolaire=annee_active,
                                    mois=mois_courant,
                                    montant=montant_mois,
                                    avance=ajout,
                                    solde=max(montant_mois - ajout, Decimal('0')),
                                    object_paiement='minerval',
                                    type_paiement='minerval',
                                    date_paiement=timezone.now().date(),
                                    recu=True,
                                    devise=devise,
                                    montant_usd_brut=montant_usd_reparti,
                                    montant_cdf_brut=montant_cdf_reparti,
                                    taux=paiement.taux,
                                    created_by=request.user
                                )

                                avance_restante -= ajout
                                paiements_crees.append(nouveau_paiement)

                        if paiements_crees and any(p.avance > 0 for p in paiements_crees):
                            return redirect('receipt_view', paiement_id=paiements_crees[0].id)

                    fm.add_error(None, "Aucun paiement n’a pu être créé. Vérifie les mois impayés ou les montants disponibles.")
                    
    classes = Classes.objects.filter(
        annee_scolaire=annee_active,
        created_by=request.user
    )

    autres_frais = AutreFrais.objects.filter(
        annee_scolaire=annee_active,
        created_by=request.user
    )

    context = {
        'paiements': paiements.filter(recu=True),
        'form': fm,
        'classes': classes,
        'autres_frais': autres_frais,
        'annee_scolaire_id': annee_active.id
    }

    return render(request, 'paiements.html', context)


@login_required
def add_class(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_scolaire:
        messages.error(request, "Année scolaire invalide ou non active.")
        return redirect('index')

    if request.method == 'POST':
        fm = ClassesRegistratation(request.POST)
        if fm.is_valid():
            new_class = fm.save(commit=False)
            new_class.annee_scolaire = annee_scolaire
            new_class.created_by = request.user  # 🔐 Séparation comptable
            new_class.save()
            messages.success(request, "Classe enregistrée avec succès.")
            return redirect('add_class')
        else:
            messages.error(request, "Erreur dans le formulaire. Veuillez corriger les champs.")
    else:
        fm = ClassesRegistratation()

    # 🔍 Affiche uniquement les classes créées par ce comptable
    classes = Classes.objects.filter(
        annee_scolaire=annee_scolaire,
        created_by=request.user
    )
    number_class = classes.count()

    context = {
        'classes': classes,
        'form': fm,
        'number_class': number_class,
        'annee_scolaire': annee_scolaire
    }
    return render(request, 'add_class.html', context)

@login_required
def UpdateClasse(request, id):
    classe = get_object_or_404(Classes, pk=id)

    # 🔐 Sécurité : seule le comptable qui a créé la classe peut la modifier
    if classe.created_by != request.user:
        return HttpResponseForbidden("Modification non autorisée")

    if request.method == 'POST':
        fm = ClassesRegistratation(request.POST, instance=classe)
        if fm.is_valid():
            fm.save()
            messages.success(request, "Classe mise à jour avec succès.")
            return redirect('add_class')
        else:
            messages.error(request, "Veuillez corriger les erreurs du formulaire.")
    else:
        fm = ClassesRegistratation(instance=classe)

    return render(request, 'add_class.html', {'form': fm, 'classe': classe})

@login_required
def DeleteClasse(request, id):
    classe = get_object_or_404(Classes, pk=id)

    if classe.created_by != request.user:
        return HttpResponseForbidden("Suppression non autorisée")

    if request.method == 'POST':
        classe.delete()
        messages.warning(request, "Classe supprimée.")
        return redirect('add_class')
    
    
@login_required
def details_classe(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    annee_active = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_active:
        messages.error(request, "Aucune année scolaire valide sélectionnée.")
        return render(request, 'details_classe.html', {
            'classes': [],
            'eleves': None,
            'selected_class_name': None
        })

    # 📚 Liste des classes créées par le comptable connecté
    classes = Classes.objects.filter(
        annee_scolaire=annee_active,
        created_by=request.user
    )

    eleves = None
    selected_class_name = None

    if request.method == 'POST':
        classe_id = request.POST.get('classe')
        if classe_id:
            try:
                classe = Classes.objects.get(
                    id=classe_id,
                    annee_scolaire=annee_active,
                    created_by=request.user  # 🔐 Sécurité comptable
                )
                eleves = student.objects.filter(
                    classe=classe,
                    annee_scolaire=annee_active,
                    created_by=request.user  # 🔍 Élèves inscrits par ce comptable uniquement
                )
                selected_class_name = classe.nom_classe
            except Classes.DoesNotExist:
                eleves = None
                selected_class_name = 'Classe introuvable'

    content = {
        'classes': classes,
        'eleves': eleves,
        'selected_class_name': selected_class_name
    }

    return render(request, 'details_classe.html', content)

    
#GESTION ANNEE SCOLAIRE PAR LE SUPER ADMIN
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from .models import AnneeScolaire
from .forms import AnneeScolaireRegistration
from django.contrib import messages

@login_required
def add_year(request):
    if request.user.role != 'superadmin':
        return HttpResponseForbidden("Accès refusé — seul le superadmin peut créer une année scolaire.")

    annees_existantes = AnneeScolaire.objects.filter(ecole__isnull=False).select_related('ecole')
    nombre_total = annees_existantes.count()

    if request.method == 'POST':
        form = AnneeScolaireRegistration(request.POST)
        if form.is_valid():
            new_year = form.save(commit=False)
            new_year.active = True  # ✅ Par défaut, l’année est active
            new_year.save()
            messages.success(request, f"Année scolaire '{new_year.nom_annee}' ajoutée avec succès.")
            return redirect('add_year')
    else:
        form = AnneeScolaireRegistration()

    context = {
        'annee_scolaire': annees_existantes,
        'form': form,
        'number_year': nombre_total
    }
    return render(request, 'add_year.html', context)

@login_required
def UpdateYear(request, id):
    if request.user.role != 'superadmin':
        return HttpResponseForbidden("Accès refusé — seul le superadmin peut modifier une année scolaire.")

    annee = get_object_or_404(AnneeScolaire, pk=id)

    if request.method == 'POST':
        form = AnneeScolaireRegistration(request.POST, instance=annee)
        if form.is_valid():
            form.save()
            messages.success(request, "Année scolaire mise à jour avec succès.")
            return redirect('add_year')
    else:
        form = AnneeScolaireRegistration(instance=annee)

    return render(request, 'updateyear.html', {'form': form, 'annee': annee})

@login_required
def DeleteYear(request, id):
    if request.user.role != 'superadmin':
        return HttpResponseForbidden("Accès refusé — seul le superadmin peut supprimer une année scolaire.")

    if request.method == 'POST':
        annee = get_object_or_404(AnneeScolaire, pk=id)
        messages.warning(request, f"Année scolaire '{annee.nom_annee}' supprimée.")
        annee.delete()
        return redirect('add_year')
    

#AUTRES FRAIS

@login_required
def ajouter_autre_frais(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès non autorisé")

    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_scolaire:
        messages.error(request, "Année scolaire active introuvable.")
        return redirect('index')

    if request.method == 'POST':
        form = AutreFraisForm(request.POST)
        if form.is_valid():
            nouveau_frais = form.save(commit=False)
            nouveau_frais.annee_scolaire = annee_scolaire
            nouveau_frais.created_by = request.user  # 🔐 Attribution comptable
            nouveau_frais.save()
            messages.success(request, "Frais enregistré avec succès.")
            return redirect('ajouter_autre_frais')
        else:
            messages.error(request, "Erreur dans le formulaire.")
    else:
        form = AutreFraisForm()

    autres_frais = AutreFrais.objects.filter(
        annee_scolaire=annee_scolaire,
        created_by=request.user
    )

    context = {
        'form': form,
        'autres_frais': autres_frais,
        'annee_scolaire': annee_scolaire
    }
    return render(request, 'add_autre_frais.html', context)

@login_required
def UpdateAutreFrais(request, id):
    frais = get_object_or_404(AutreFrais, pk=id)

    if request.user.role != 'comptable' or frais.created_by != request.user:
        return HttpResponseForbidden("Modification non autorisée")

    annee_id = request.session.get('annee_scolaire')
    if not annee_id or frais.annee_scolaire.id != int(annee_id):
        messages.error(request, "Année scolaire incorrecte.")
        return redirect('ajouter_autre_frais')

    if request.method == 'POST':
        fm = AutreFraisForm(request.POST, instance=frais)
        if fm.is_valid():
            fm.save()
            messages.success(request, "Frais mis à jour.")
            return redirect('ajouter_autre_frais')
        else:
            messages.error(request, "Erreur de validation.")
    else:
        fm = AutreFraisForm(instance=frais)

    return render(request, 'add_autre_frais.html', {'form': fm, 'frais': frais})

@login_required
def DeleteAutreFrais(request, id):
    frais = get_object_or_404(AutreFrais, pk=id)

    if request.user.role != 'comptable' or frais.created_by != request.user:
        return HttpResponseForbidden("Suppression non autorisée")

    if request.method == 'POST':
        frais.delete()
        messages.warning(request, "Frais supprimé.")
        return redirect('ajouter_autre_frais')

    return render(request, 'confirm_delete_frais.html', {'frais': frais})

#LES RAPPORTS JOUR-MOIS-ANNEE
@login_required
def rapport_paiements_jour(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès interdit.")

    # 📅 Lecture de la date à afficher
    date_str = request.GET.get('date', now().date().strftime('%Y-%m-%d'))
    try:
        date_selected = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        date_selected = now().date()

    # 📚 Récupération de l'année scolaire active
    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_scolaire:
        messages.error(request, "Aucune année scolaire active.")
        return redirect('logout')

    # 📥 Paiements et dépenses du jour
    paiements = Paiement.objects.filter(
        date_paiement=date_selected,
        annee_scolaire=annee_scolaire,
        recu=True,
        created_by=request.user
    )

    depenses = Depense.objects.filter(
        date_depense=date_selected,
        annee_scolaire=annee_scolaire,
        created_by=request.user
    )

    # 💰 Totaux encaissés du jour — séparés sans doublon ni conversion
    montant_cdf = paiements.aggregate(
        total=Sum('montant_cdf_brut')
    )['total'] or Decimal('0')

    montant_usd = paiements.aggregate(
        total=Sum('montant_usd_brut')
    )['total'] or Decimal('0')

    # 🧾 Dépenses du jour
    depense_cdf = depenses.aggregate(
        total=Sum('montant')
    )['total'] or Decimal('0')

    depense_usd = depenses.aggregate(
        total=Sum('montant_usd')
    )['total'] or Decimal('0')

    context = {
        'date_selected': date_selected,
        'paiements': paiements,
        'depenses': depenses,
        'montant_cdf': montant_cdf,
        'montant_usd': montant_usd,
        'depense_cdf': depense_cdf,
        'depense_usd': depense_usd,
    }

    return render(request, 'rapport_paiements_jour.html', context)

from urllib.parse import urlencode

@login_required
def rapport_mensuel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès interdit.")

    mois_selectionne = None
    classe_selectionnee = None
    paiements = []
    montant_total_cdf = Decimal('0')
    montant_total_usd = Decimal('0')

    annee_id = request.session.get('annee_scolaire')
    annee_actuelle = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if request.method == "POST":
        form = MoisClasseFilterForm(request.POST, annee=annee_actuelle)
        # 🛠️ Ajustement ici : affichage des classes du comptable
        form.fields['classe'].choices = [('', 'Toutes les classes')] + [
            (c.id, c.nom_classe) for c in Classes.objects.filter(
                annee_scolaire=annee_actuelle,
                created_by=request.user
            )
        ]

        if form.is_valid():
            mois_selectionne = form.cleaned_data['mois']
            classe_selectionnee = form.cleaned_data['classe']

            filtre_base = {
                'date_paiement__month': mois_selectionne,
                'annee_scolaire': annee_actuelle,
                'recu': True,
                'created_by': request.user
            }

            if classe_selectionnee == 'all_classes':
                paiements = Paiement.objects.filter(**filtre_base)
            else:
                paiements = Paiement.objects.filter(**filtre_base, classe_id=classe_selectionnee)

            montant_total_cdf = paiements.aggregate(
                total=Sum('montant_cdf_brut')
            )['total'] or Decimal('0')

            montant_total_usd = paiements.aggregate(
                total=Sum('montant_usd_brut')
            )['total'] or Decimal('0')

            query_params = {'mois': mois_selectionne}
            if classe_selectionnee and classe_selectionnee != 'all_classes':
                query_params['classe'] = classe_selectionnee
            export_url = f"{request.build_absolute_uri('/export-rapport-mensuel/')}?{urlencode(query_params)}"
    else:
        form = MoisClasseFilterForm(annee=annee_actuelle)
        form.fields['classe'].choices = [('', 'Toutes les classes')] + [
            (c.id, c.nom_classe) for c in Classes.objects.filter(
                annee_scolaire=annee_actuelle,
                created_by=request.user
            )
        ]
        export_url = None

    context = {
        'form': form,
        'paiements': paiements,
        'mois_selectionne': mois_selectionne,
        'classe_selectionnee': classe_selectionnee,
        'export_url': export_url,
        'annee_actuelle': annee_actuelle,
        'montant_total_cdf': montant_total_cdf,
        'montant_total_usd': montant_total_usd,
    }
    return render(request, 'rapport_mensuel.html', context)

@login_required
def rapport_annuel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès interdit.")

    annee_id = request.session.get('annee_scolaire')
    annee_actuelle = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    classe_selectionnee = None
    paiements = []
    montant_total_cdf = Decimal('0')
    montant_total_usd = Decimal('0')

    if request.method == "POST":
        form = ClasseFilterForm(request.POST, annee=annee_actuelle)
        form.fields['classe'].queryset = Classes.objects.filter(
            annee_scolaire=annee_actuelle,
            created_by=request.user
        )
        if form.is_valid():
            classe_selectionnee = form.cleaned_data['classe']

            filtre_base = {
                'classe__annee_scolaire': annee_actuelle,
                'recu': True,
                'created_by': request.user
            }

            if classe_selectionnee:
                paiements = Paiement.objects.filter(**filtre_base, classe=classe_selectionnee)
            else:
                paiements = Paiement.objects.filter(**filtre_base)

            montant_total_cdf = paiements.aggregate(
                total=Sum('montant_cdf_brut')
            )['total'] or Decimal('0')

            montant_total_usd = paiements.aggregate(
                total=Sum('montant_usd_brut')
            )['total'] or Decimal('0')
    else:
        form = ClasseFilterForm(annee=annee_actuelle)
        form.fields['classe'].queryset = Classes.objects.filter(
            annee_scolaire=annee_actuelle,
            created_by=request.user
        )

    context = {
        'form': form,
        'paiements': paiements,
        'montant_total_cdf': montant_total_cdf,
        'montant_total_usd': montant_total_usd,
        'classe_selectionnee': classe_selectionnee,
        'annee_actuelle': annee_actuelle,
    }
    return render(request, 'rapport_annuel.html', context)

#VUE IMPRESSION 
from django.utils.timezone import now
from io import BytesIO
import base64
import qrcode

@login_required
def receipt_view(request, paiement_id):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès non autorisé")

    paiement = get_object_or_404(Paiement, id=paiement_id)

    # 🔐 Vérifie que ce paiement appartient au comptable
    if paiement.created_by != request.user:
        return HttpResponseForbidden("Ce reçu ne vous appartient pas.")

    eleve = paiement.eleve

    devise_label = "CDF" if paiement.devise == "cdf" else "USD"
    montant_affiche = f"{paiement.montant} {devise_label}"

    # 🧾 Génération du QR code
    data_to_encode = f"{eleve.nom} | {eleve.classe.nom_classe} | {montant_affiche}"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=0
    )
    qr.add_data(data_to_encode)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    img_io = BytesIO()
    img.save(img_io)
    img_io.seek(0)
    qr_base64 = base64.b64encode(img_io.read()).decode('utf-8')

    # 📄 Sélection du bon template
    if paiement.type_paiement == 'minerval':
        template_name = 'receipt_multi.html'
        paiements_affiches = Paiement.objects.filter(
            eleve=eleve,
            annee_scolaire=paiement.annee_scolaire,
            date_paiement=paiement.date_paiement,
            type_paiement='minerval',
            created_by=request.user
        ).order_by('mois')
    else:
        template_name = 'receipt.html'
        paiements_affiches = [paiement]

    context = {
        'paiement': paiement,
        'paiements': paiements_affiches,
        'qr_code_image': qr_base64,
        'montant_affiche': montant_affiche,
        'devise': devise_label
    }

    return render(request, template_name, context)

#DETAILS PAIEMENT COMPTABLE

@login_required
def details_paiements(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    paiements = []

    # 🔍 Vérification de l'année scolaire active
    annee_active = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_active:
        messages.error(request, "Année scolaire invalide.")
        return redirect('index')

    # 📚 Récupération des classes de l'utilisateur
    classes = Classes.objects.filter(
        annee_scolaire=annee_active,
        created_by=request.user
    )

    # 🧠 Initialisation du formulaire avec contexte
    form = PaiementRegistration(
        request.POST or None,
        annee_scolaire_id=annee_active.id,
        created_by=request.user
    )

    if request.method == 'POST':
        if form.is_valid():
            paiement = form.save(commit=False)
            paiement.annee_scolaire = annee_active
            paiement.created_by = request.user
            paiement.save()
            messages.success(request, 'Le paiement a été enregistré avec succès.')
            # 🔄 Réinitialiser le formulaire après enregistrement
            form = PaiementRegistration(
                annee_scolaire_id=annee_active.id,
                created_by=request.user
            )

        # 🎯 Filtrage dynamique des paiements
        classe = request.POST.get('classe')
        mois = request.POST.get('mois')
        object_paiement = request.POST.get('object_paiement')
        recu = request.POST.get('recu')

        paiements = Paiement.objects.filter(
            annee_scolaire=annee_active,
            created_by=request.user
        )
        if classe:
            paiements = paiements.filter(classe_id=classe)
        if mois:
            paiements = paiements.filter(mois=mois)
        if object_paiement:
            paiements = paiements.filter(object_paiement=object_paiement)
        if recu == "True":
            paiements = paiements.filter(recu=True)
        elif recu == "False":
            paiements = paiements.filter(recu=False)

    context = {
        'paiements': paiements,
        'form': form,
        'classes': classes,
        'annee_scolaire_id': annee_active.id,
    }
    return render(request, 'details_paiements.html', context)





@login_required
def eleves_non_payeurs(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    eleves_non_payeurs = student.objects.none()

    annee_active = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_active:
        messages.error(request, "Année scolaire invalide.")
        return redirect('index')

    classes = Classes.objects.filter(
        annee_scolaire=annee_active,
        created_by=request.user
    )

    # 🧠 Initialisation du formulaire avec contexte
    form = PaiementRegistration(
        request.POST or None,
        annee_scolaire_id=annee_active.id,
        created_by=request.user
    )

    mois = object_paiement = None

    if request.method == 'POST':
        if form.is_valid():
            paiement = form.save(commit=False)
            paiement.annee_scolaire = annee_active
            paiement.created_by = request.user
            paiement.save()
            messages.success(request, 'Paiement enregistré.')
            form = PaiementRegistration(
                annee_scolaire_id=annee_active.id,
                created_by=request.user
            )

        classe_id = request.POST.get('classe')
        mois = request.POST.get('mois')
        object_paiement = request.POST.get('object_paiement')

        if classe_id and mois and object_paiement:
            eleves_classe = student.objects.filter(
                classe_id=classe_id,
                annee_scolaire=annee_active,
                created_by=request.user
            )

            payeurs = Paiement.objects.filter(
                classe_id=classe_id,
                mois=mois,
                object_paiement=object_paiement,
                recu=True,
                created_by=request.user
            ).values_list('eleve_id', flat=True)

            eleves_non_payeurs = eleves_classe.exclude(id__in=payeurs)

    context = {
        'eleves_non_payeurs': eleves_non_payeurs,
        'form': form,
        'mois': mois or '',
        'object_paiement': object_paiement or '',
        'classes': classes,
    }
    return render(request, 'notifications.html', context)




#FONCTIONS POUR LA SELECTION DES ELEVES A PARTIR D'UNE CLASSE
# 🔹 Minerval d’une classe
@login_required
def get_montant_minerval(request, classe_id):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès non autorisé")

    try:
        classe = Classes.objects.get(
            id=classe_id,
            created_by=request.user
        )
        return JsonResponse({'montant_minerval': float(classe.montant)})
    except Classes.DoesNotExist:
        return JsonResponse({'error': 'Classe non trouvée'}, status=404)

# 🔹 Recherche dynamique d’élèves par nom
@login_required
def search_students(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    query = request.GET.get('query', '')
    classe_id = request.GET.get('classe_id')

    students = student.objects.filter(
        classe_id=classe_id,
        nom__icontains=query,
        created_by=request.user
    ) if classe_id else student.objects.none()

    results = [{"id": st.id, "name": f"{st.nom} {st.post_nom} {st.prenom}"} for st in students]
    return JsonResponse(results, safe=False)

# 🔹 Liste d’élèves par classe via Ajax
@login_required
def get_eleves(request, classe_id):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    if request.method == 'GET' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        eleves = student.objects.filter(
            classe_id=classe_id,
            created_by=request.user
        )
        data = [{'id': e.id, 'nom': f"{e.nom} {e.prenom}"} for e in eleves]
        return JsonResponse(data, safe=False)

    return JsonResponse({'error': 'Requête invalide'}, status=400)

# 🔹 Élèves via GET simple
@login_required
def load_eleves(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    classe_id = request.GET.get('classe')
    eleves = student.objects.filter(
        classe_id=classe_id,
        created_by=request.user
    ).values('id', 'nom', 'prenom')
    return JsonResponse(list(eleves), safe=False)

# 🔹 Classes par année (côté comptable)
@login_required
def get_classes_by_annee(request, annee_id):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    classes = Classes.objects.filter(
        annee_scolaire_id=annee_id,
        created_by=request.user
    ).values('id', 'nom_classe')

    return JsonResponse(list(classes), safe=False)

# 🔹 Élèves d’une classe avec gestion d’erreurs
@login_required
def get_eleves_by_classe(request, classe_id):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    try:
        classe = get_object_or_404(
            Classes,
            id=int(classe_id),
            created_by=request.user
        )

        eleves = student.objects.filter(
            classe=classe,
            created_by=request.user
        ).values('id', 'nom', 'post_nom', 'prenom')

        if not eleves.exists():
            return JsonResponse({'error': 'Aucun élève trouvé pour cette classe'}, status=404)

        return JsonResponse(list(eleves), safe=False)

    except ValueError:
        return JsonResponse({'error': 'Identifiant invalide'}, status=400)
    except DatabaseError as e:
        print(f"Erreur DB : {str(e)}")
        return JsonResponse({'error': 'Erreur de base de données'}, status=500)
    except Exception as e:
        print(f"Erreur inattendue : {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)
    
#API POUR CHARGER LES AUTRES FRAIS

@login_required
def get_autre_frais_options(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    devise = request.GET.get('devise', 'cdf')
    annee_id = request.session.get('annee_scolaire')

    annee_scolaire = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_scolaire:
        return JsonResponse({'error': 'Aucune année active'}, status=400)

    frais = AutreFrais.objects.filter(
        annee_scolaire=annee_scolaire,
        created_by=request.user  # 🔒 Isolation par comptable
    )

    data = []
    for f in frais:
        montant = f.montant_cdf if devise == 'cdf' else f.montant_usd
        data.append({
            'id': f.id,
            'description': f.description,
            'montant': float(montant),
        })

    return JsonResponse({'frais': data})
    
#FONCTIONS POUR L'EXPORTATION AU FORMAT EXCEL
import xlwt

@login_required
def export_paiements_excel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    filtered_paiements = request.session.get('filtered_paiements', [])

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=paiements.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Paiements')

    columns = ['Nom', 'Post Nom', 'Prénom', 'Classe', 'Montant', 'Mois', 'Objet du paiement', 'Date de paiement']
    for col_num, col_title in enumerate(columns):
        ws.write(0, col_num, col_title)

    for row_num, paiement in enumerate(filtered_paiements, start=1):
        ws.write(row_num, 0, paiement['eleve__nom'])
        ws.write(row_num, 1, paiement['eleve__post_nom'])
        ws.write(row_num, 2, paiement['eleve__prenom'])
        ws.write(row_num, 3, paiement['classe__nom_classe'])
        ws.write(row_num, 4, paiement['montant'])
        ws.write(row_num, 5, paiement['mois'])
        ws.write(row_num, 6, paiement['object_paiement'])
        ws.write(row_num, 7, paiement['date_paiement'])

    wb.save(response)
    return response

@login_required
def export_insolvable_excel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    classe_id = request.GET.get('classe')
    mois = request.GET.get('mois')
    object_paiement = request.GET.get('object_paiement')

    if not (classe_id and mois and object_paiement):
        return HttpResponse("Paramètres manquants.", status=400)

    eleves_classe = student.objects.filter(
        classe_id=classe_id,
        annee_scolaire_id=annee_id,
        created_by=request.user  # 🔒 Isolation comptable
    )

    payeurs = Paiement.objects.filter(
        classe_id=classe_id,
        mois=mois,
        object_paiement=object_paiement,
        recu=True,
        created_by=request.user  # 🔐 Paiements du comptable courant uniquement
    ).values_list('eleve_id', flat=True)

    eleves_non_payeurs = eleves_classe.exclude(id__in=payeurs)

    if not eleves_non_payeurs.exists():
        return HttpResponse("Aucun élève non payeur trouvé.", status=404)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=eleves_non_payeurs.xls'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Elèves Non Payeurs')

    headers = ['Nom Élève', 'Classe', 'Mois', 'Objet Paiement']
    for col_num, header in enumerate(headers):
        ws.write(0, col_num, header)

    for row_num, eleve in enumerate(eleves_non_payeurs, start=1):
        ws.write(row_num, 0, f"{eleve.nom} {eleve.post_nom} {eleve.prenom}")
        ws.write(row_num, 1, eleve.classe.nom_classe)
        ws.write(row_num, 2, mois)
        ws.write(row_num, 3, object_paiement)

    wb.save(response)
    return response


from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.db.models import Case, When, F, Sum
from .models import student, Classes, Paiement, AnneeScolaire
from datetime import datetime
import xlwt

# 🔹 1. Exporter les élèves d’une classe
@login_required
def export_eleves_excel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    classe_id = request.GET.get('classe')

    selected_class_name = 'Inconnue'
    eleves = student.objects.none()

    if classe_id and annee_id:
        try:
            classe = Classes.objects.get(id=classe_id, created_by=request.user)
            eleves = student.objects.filter(classe=classe, annee_scolaire_id=annee_id, created_by=request.user)
            selected_class_name = classe.nom_classe
        except Classes.DoesNotExist:
            pass

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="eleves_classe_{selected_class_name}.xls"'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet(f"Classe {selected_class_name}")

    headers = ['Nom', 'Post_nom', 'Prenom', 'Sexe', 'Classe', 'Téléphone']
    for col, header in enumerate(headers):
        ws.write(0, col, header)

    for row, eleve in enumerate(eleves, start=1):
        ws.write(row, 0, eleve.nom)
        ws.write(row, 1, eleve.post_nom)
        ws.write(row, 2, eleve.prenom)
        ws.write(row, 3, eleve.sexe)
        ws.write(row, 4, eleve.classe.nom_classe)
        ws.write(row, 5, eleve.telephone)

    wb.save(response)
    return response

# 🔹 2. Rapport journalier
@login_required
def export_rapport_journalier_excel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    date_str = request.GET.get('date')

    if not annee_id or not date_str:
        return HttpResponse("Année ou date manquante.", status=400)

    try:
        date_selected = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse("Format de date invalide.", status=400)

    paiements = Paiement.objects.filter(
        date_paiement=date_selected,
        eleve__annee_scolaire_id=annee_id,
        recu=True,
        created_by=request.user
    )

    if not paiements.exists():
        return HttpResponse("Aucun paiement trouvé.", status=404)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=rapport_journalier_{date_str}.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Rapport Journalier')

    headers = ['Nom Élève', 'Classe', 'Montant Payé', 'Date Paiement', 'Objet Paiement']
    for col, header in enumerate(headers):
        ws.write(0, col, header, xlwt.easyxf('font: bold on'))

    for row, p in enumerate(paiements, start=1):
        ws.write(row, 0, f"{p.eleve.nom} {p.eleve.post_nom} {p.eleve.prenom}")
        ws.write(row, 1, p.classe.nom_classe)
        ws.write(row, 2, p.montant)
        ws.write(row, 3, p.date_paiement.strftime('%d/%m/%Y'))
        ws.write(row, 4, p.object_paiement)

    wb.save(response)
    return response

# 🔹 3. Rapport mensuel
@login_required
def export_rapport_mensuel_excel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    mois = request.GET.get('mois')
    classe_id = request.GET.get('classe')
    annee_id = request.session.get('annee_scolaire')

    if not mois or not annee_id:
        return HttpResponse("Paramètres manquants.", status=400)

    paiements = Paiement.objects.filter(
        date_paiement__month=mois,
        annee_scolaire_id=annee_id,
        recu=True,
        created_by=request.user
    )

    if classe_id:
        paiements = paiements.filter(classe_id=classe_id)

    if not paiements.exists():
        return HttpResponse("Aucun paiement trouvé.", status=404)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=rapport_mensuel_{mois}.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Rapport Mensuel')

    headers = ['Nom Élève', 'Classe', 'Montant', 'Mois', 'Objet', 'Date']
    for col, header in enumerate(headers):
        ws.write(0, col, header, xlwt.easyxf('font: bold on'))

    for row, p in enumerate(paiements, start=1):
        ws.write(row, 0, f"{p.eleve.nom} {p.eleve.post_nom} {p.eleve.prenom}")
        ws.write(row, 1, p.classe.nom_classe)
        ws.write(row, 2, p.montant)
        ws.write(row, 3, mois)
        ws.write(row, 4, p.object_paiement)
        ws.write(row, 5, p.date_paiement.strftime('%d/%m/%Y'))

    wb.save(response)
    return response

# 🔹 4. Rapport annuel
@login_required
def export_rapport_annuel_excel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    classe_id = request.GET.get('classe')

    if not annee_id:
        return HttpResponse("Année scolaire manquante.", status=400)

    annee = get_object_or_404(AnneeScolaire, id=annee_id)

    paiements = Paiement.objects.filter(
        eleve__annee_scolaire=annee,
        created_by=request.user
    )

    if classe_id:
        paiements = paiements.filter(classe_id=classe_id)

    paiements = paiements.annotate(
        montant_effectif=Case(
            When(solde=0, then=F('montant')),
            default=F('avance')
        )
    )

    montant_total = paiements.aggregate(total=Sum('montant_effectif'))['total'] or 0

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename=rapport_annuel_{annee.nom_annee}.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Rapport Annuel')

    headers = ['Nom Élève', 'Classe', 'Montant', 'Mois', 'Objet', 'Date']
    for col, header in enumerate(headers):
        ws.write(0, col, header, xlwt.easyxf('font: bold on'))

    row = 1
    for p in paiements:
        ws.write(row, 0, f"{p.eleve.nom} {p.eleve.post_nom} {p.eleve.prenom}")
        ws.write(row, 1, p.classe.nom_classe)
        ws.write(row, 2, p.montant_effectif)
        ws.write(row, 3, p.date_paiement.strftime('%m/%Y'))
        ws.write(row, 4, p.object_paiement)
        ws.write(row, 5, p.date_paiement.strftime('%d/%m/%Y'))
        row += 1

    ws.write(row, 1, "Total", xlwt.easyxf('font: bold on'))
    ws.write(row, 2, montant_total, xlwt.easyxf('font: bold on'))

    wb.save(response)
    return response


#AJOUT DE RAPPORT FINANCIER INTELLIGENT 

from django.shortcuts import render, redirect
from .models import Depense, AnneeScolaire
from .forms import DepenseForm
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.shortcuts import render, redirect
from decimal import Decimal
from django.db.models import Sum
import pandas as pd
from .forms import DepenseForm
from .models import AnneeScolaire, Paiement, Depense, student, Classes

# 🔹 Vue 1 : Ajouter une dépense (comptable uniquement)
@login_required
def ajouter_depense(request):
    # 🔐 Vérifie que l'utilisateur est bien un comptable
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    # 📅 Récupère l'année scolaire active depuis la session
    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    if not annee_scolaire:
        return redirect('index')

    # 📝 Traitement du formulaire
    if request.method == 'POST':
        form = DepenseForm(
            request.POST,
            created_by=request.user,
            annee_scolaire=annee_scolaire
        )
        if form.is_valid():
            depense = form.save(commit=False)
            depense.annee_scolaire = annee_scolaire
            depense.created_by = request.user
            depense.montant = depense.montant or Decimal('0')
            depense.montant_usd = depense.montant_usd or Decimal('0')
            depense.save()
            return redirect('rapport_financier')
    else:
        form = DepenseForm(
            created_by=request.user,
            annee_scolaire=annee_scolaire
        )

    return render(request, 'ajouter_depense.html', {
        'form': form,
        'annee_scolaire': annee_scolaire
    })


# 🔹 Vue 2 : Rapport financier (comptable uniquement)
@login_required
def rapport_financier(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = AnneeScolaire.objects.filter(
        id=annee_id,
        ecole=request.user.ecole,
        active=True
    ).first()

    total_revenu = total_revenu_usd = total_depenses = total_depenses_usd = solde_net = solde_net_usd = Decimal('0')
    eleves_non_payeurs = []
    total_dette = Decimal('0')
    alertes = []

    if annee_scolaire:
        # 💰 Encaissements réels sans conversion USD
        total_revenu = Paiement.objects.filter(
            annee_scolaire=annee_scolaire,
            recu=True,
            created_by=request.user
        ).aggregate(total=Sum('montant_cdf_brut'))['total'] or Decimal('0')

        total_revenu_usd = Paiement.objects.filter(
            annee_scolaire=annee_scolaire,
            recu=True,
            created_by=request.user
        ).aggregate(total=Sum('montant_usd_brut'))['total'] or Decimal('0')

        # 🧾 Dépenses
        total_depenses = Depense.objects.filter(
            annee_scolaire=annee_scolaire,
            created_by=request.user
        ).aggregate(total=Sum('montant'))['total'] or Decimal('0')

        total_depenses_usd = Depense.objects.filter(
            annee_scolaire=annee_scolaire,
            created_by=request.user
        ).aggregate(total=Sum('montant_usd'))['total'] or Decimal('0')

        # 📊 Soldes nets
        solde_net = total_revenu - total_depenses
        solde_net_usd = total_revenu_usd - total_depenses_usd

        # 👨‍🎓 Élèves non à jour
        classes = Classes.objects.filter(
            annee_scolaire=annee_scolaire,
            created_by=request.user
        )
        for classe in classes:
            minerval_cdf = classe.montant
            eleves_classe = student.objects.filter(classe=classe, created_by=request.user)
            for eleve in eleves_classe:
                mois_payes = Paiement.objects.filter(
                    eleve=eleve,
                    type_paiement='minerval',
                    recu=True,
                    devise='cdf',
                    created_by=request.user
                ).values_list('mois', flat=True)

                mois_non_payes = [mois for mois in [
                    'septembre', 'octobre', 'novembre', 'décembre',
                    'janvier', 'février', 'mars', 'avril', 'mai', 'juin'
                ] if mois not in mois_payes]

                montant_du = len(mois_non_payes) * minerval_cdf
                if montant_du > 0:
                    eleves_non_payeurs.append({'eleve': eleve, 'montant_du': montant_du})
                    total_dette += montant_du

        # 📅 Répartition par mois et classe
        paiements_par_mois = Paiement.objects.filter(
            annee_scolaire=annee_scolaire,
            recu=True,
            created_by=request.user
        ).values('mois').annotate(total=Sum('montant_cdf_brut'))

        paiements_par_classe = Paiement.objects.filter(
            annee_scolaire=annee_scolaire,
            recu=True,
            created_by=request.user
        ).values('classe__nom_classe').annotate(total=Sum('montant_cdf_brut'))

        # ⚠️ Alertes de gestion
        if total_depenses > total_revenu:
            alertes.append("❗ Dépenses élevées : les dépenses dépassent les revenus CDF.")
        elif total_depenses < Decimal('0.7') * total_revenu:
            alertes.append("✅ Bonne gestion : les dépenses CDF représentent moins de 70% des revenus.")

        if total_depenses_usd > total_revenu_usd:
            alertes.append("❗ Dépenses USD élevées : elles dépassent les revenus USD.")
        elif total_depenses_usd < Decimal('0.7') * total_revenu_usd:
            alertes.append("✅ Bonne gestion USD : dépenses inférieures à 70% des revenus USD.")

        # 🔍 Audit des postes
        depenses_qs = Depense.objects.filter(
            annee_scolaire=annee_scolaire,
            created_by=request.user
        ).values('description', 'montant')

        if depenses_qs.exists():
            df_dep = pd.DataFrame.from_records(depenses_qs)
            grouped = df_dep.groupby('description')['montant'].sum().reset_index()
            for _, row in grouped.iterrows():
                if row['montant'] > 200000:
                    alertes.append(
                        f"🔧 Dépense élevée pour « {row['description']} » : {row['montant']} FC. Réduire de 10%."
                    )

    context = {
        'annee_scolaire': annee_scolaire,
        'total_revenu': total_revenu,
        'total_revenu_usd': total_revenu_usd,
        'total_depenses': total_depenses,
        'total_depenses_usd': total_depenses_usd,
        'solde_net': solde_net,
        'solde_net_usd': solde_net_usd,
        'depenses': Depense.objects.filter(annee_scolaire=annee_scolaire, created_by=request.user),
        'eleves_non_payeurs': eleves_non_payeurs,
        'total_dette': total_dette,
        'paiements_par_mois': paiements_par_mois,
        'paiements_par_classe': paiements_par_classe,
        'alertes': alertes,
    }

    return render(request, 'rapport_financier.html', context)

import xlwt
from django.http import HttpResponse
from reportlab.pdfgen import canvas

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render, get_object_or_404
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, PageBreak
from .models import AnneeScolaire, Depense, student, Paiement, Classes
import xlwt
import os
from decimal import Decimal


# 🔹 1. Export Excel - Rapport financier global (comptable uniquement)
@login_required
def export_rapport_excel(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = get_object_or_404(AnneeScolaire, id=annee_id)

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=rapport_financier.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Rapport Financier')

    ws.write(0, 0, 'Total Revenu')
    ws.write(0, 1, float(annee_scolaire.total_revenu or 0))

    ws.write(1, 0, 'Total Dépenses')
    ws.write(1, 1, float(annee_scolaire.total_depenses or 0))

    ws.write(2, 0, 'Solde Net')
    ws.write(2, 1, float((annee_scolaire.total_revenu or 0) - (annee_scolaire.total_depenses or 0)))

    wb.save(response)
    return response

# 🔹 2. Export PDF - Rapport financier global (comptable uniquement)
@login_required
def export_rapport_pdf(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = get_object_or_404(AnneeScolaire, id=annee_id)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=rapport_financier.pdf'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    titre = Paragraph(f"<b>Rapport Financier - {annee_scolaire.nom_annee}</b>", styles["Title"])
    description = Paragraph(
        f"Le rapport pour l'année scolaire {annee_scolaire.nom_annee} synthétise les revenus encaissés, "
        f"les dépenses enregistrées, et le solde net constaté.",
        styles["BodyText"]
    )

    table_data = [
        ["Catégorie", "Montant (CDF)"],
        ["Total Revenu", f"{annee_scolaire.total_revenu or 0} FC"],
        ["Total Dépenses", f"{annee_scolaire.total_depenses or 0} FC"],
        ["Solde Net", f"{(annee_scolaire.total_revenu or 0) - (annee_scolaire.total_depenses or 0)} FC"]
    ]

    table = Table(table_data, colWidths=[220, 180])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#F3F4F6")),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    footer = Paragraph(
        f"Rapport généré pour l'année scolaire {annee_scolaire.nom_annee} | Établissement XYZ",
        styles["Normal"]
    )

    elements.extend([titre, description, Spacer(1, 20), table, PageBreak(), footer])
    doc.build(elements)
    return response

# 🔹 3. Affichage des dépenses
from .forms import FiltreDepenseForm
from datetime import datetime

@login_required
def voir_toutes_les_depenses(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    # Récupération de l'année scolaire en session
    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = get_object_or_404(AnneeScolaire, id=annee_id)

    # Initialisation du formulaire
    form = FiltreDepenseForm(request.POST or None)

    # Dépenses liées à l'année scolaire et à l'utilisateur
    depenses = Depense.objects.filter(
        annee_scolaire=annee_scolaire,
        created_by=request.user
    )

    # Variables pour le résumé
    mois_nom = ''
    description_nom = ''

    # Filtrage si formulaire soumis
    if request.method == 'POST' and form.is_valid():
        mois = form.cleaned_data.get('mois')
        description = form.cleaned_data.get('description')

        if mois:
            depenses = depenses.filter(date_depense__month=int(mois))
            mois_nom = dict(form.fields['mois'].choices).get(mois, '')

        if description:
            depenses = depenses.filter(description=description)
            description_nom = description.description # Affiche le nom lisible

    # Calcul des totaux
    montant_total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    montant_total_depenses_usd = depenses.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0')

    # Contexte pour le template
    context = {
        'form': form,
        'depenses': depenses,
        'montant_total_depenses': montant_total_depenses,
        'montant_total_depenses_usd': montant_total_depenses_usd,
        'mois_nom': mois_nom,
        'description_nom': description_nom
    }

    return render(request, 'toutes_les_depenses.html', context)


#exporter depense pdf
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa  # Assure-toi que xhtml2pdf est installé

@login_required
def exporter_depenses_pdf(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = get_object_or_404(AnneeScolaire, id=annee_id)

    depenses = Depense.objects.filter(annee_scolaire=annee_scolaire, created_by=request.user)
    total_cdf = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    total_usd = depenses.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0')

    context = {
        'depenses': depenses,
        'total_cdf': total_cdf,
        'total_usd': total_usd
    }

    template = get_template('pdf_depenses.html')  # On crée ce template juste après
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="depenses.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

# 🔹 4. Détail des dettes (par classe)
@login_required
def dettes_details(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    annee_id = request.session.get('annee_scolaire')
    annee_scolaire = get_object_or_404(AnneeScolaire, id=annee_id)

    total_dette = 0
    eleves_non_payeurs = []
    classes_concernees = []

    classes = Classes.objects.filter(annee_scolaire=annee_scolaire, created_by=request.user)
    for classe in classes:
        montant_mensuel = classe.montant
        eleves_classe = student.objects.filter(classe=classe, created_by=request.user)

        for eleve in eleves_classe:
            mois_payes = Paiement.objects.filter(
                eleve=eleve, type_paiement='minerval', recu=True, created_by=request.user
            ).values_list('mois', flat=True)
            mois_non_payes = [
                mois for mois in [
                    'septembre', 'octobre', 'novembre', 'décembre',
                    'janvier', 'février', 'mars', 'avril', 'mai', 'juin'
                ] if mois not in mois_payes
            ]

            montant_du = len(mois_non_payes) * montant_mensuel
            if montant_du > 0:
                eleves_non_payeurs.append({'eleve': eleve, 'montant_du': montant_du})
                total_dette += montant_du
                if classe not in classes_concernees:
                    classes_concernees.append(classe)

    selected_classe = request.GET.get('classe')
    eleves_filtrés = []

    if selected_classe:
        classe_selectionnee = get_object_or_404(Classes, id=selected_classe)
        for item in eleves_non_payeurs:
            if item['eleve'].classe.id == classe_selectionnee.id:
                eleves_filtrés.append(item)
    else:
        eleves_filtrés = eleves_non_payeurs

    context = {
        'classes_concernees': classes_concernees,
        'selected_classe': selected_classe,
        'eleves_non_payeurs': eleves_filtrés,
        'total_dette': total_dette,
    }
    return render(request, 'dettes_details.html', context)

# 🔹 5. Détail dette élève individuel
@login_required
def details_eleve_dette(request, eleve_id):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    eleve = get_object_or_404(student, id=eleve_id, created_by=request.user)
    classe = eleve.classe
    montant_mensuel = classe.montant

    mois_payes = Paiement.objects.filter(
        eleve=eleve, type_paiement='minerval', recu=True, created_by=request.user
    ).values_list('mois', flat=True)
    mois_non_payes = [
        mois for mois in [
            'septembre', 'octobre', 'novembre', 'décembre',
            'janvier', 'février', 'mars', 'avril', 'mai', 'juin'
        ] if mois not in mois_payes
    ]

    details_dettere = [{'mois': mois, 'montant': montant_mensuel} for mois in mois_non_payes]

    context = {
        'eleve': eleve,
        'classe': classe,
        'details_dettere': details_dettere,
        'montant_total_du': sum([d['montant'] for d in details_dettere])
    }
    return render(request, 'details_eleve_dette.html', context)

# 🔹 6. Élèves en retard par classe
@login_required
def eleves_en_retard(request):
    if request.user.role != 'comptable':
        return HttpResponseForbidden("Accès refusé")

    classe_id = request.GET.get('classe')
    classe = get_object_or_404(Classes, id=classe_id, created_by=request.user)
    montant_mensuel = classe.montant

    eleves_classe = student.objects.filter(classe=classe, created_by=request.user)
    eleves_non_payeurs = []

    for eleve in eleves_classe:
        mois_payes = Paiement.objects.filter(
            eleve=eleve, type_paiement='Paiement', recu=True, created_by=request.user
        ).values_list('mois', flat=True)
        mois_non_payes = [
            mois for mois in [
                'septembre', 'octobre', 'novembre', 'décembre',
                'janvier', 'février', 'mars', 'avril', 'mai', 'juin'
            ] if mois not in mois_payes
        ]
        montant_du = len(mois_non_payes) * montant_mensuel
        if montant_du > 0:
            eleves_non_payeurs.append({'eleve': eleve, 'montant_du': montant_du})

    context = {
        'classe': classe,
        'eleves_non_payeurs': eleves_non_payeurs
    }
    return render(request, 'eleves_en_retard.html', context)


#COTE PROMO VUES PROMO 
@login_required
def liste_comptables(request):
    if request.user.role != 'promo':
        return HttpResponseForbidden("Accès interdit")

    comptables = User.objects.filter(
        role='comptable',
        ecole=request.user.ecole
    ).order_by('-date_joined')

    search_query = request.GET.get('search')
    if search_query:
        comptables = comptables.filter(nom_complet__icontains=search_query)

    context = {
        'comptables': comptables,
        'search_query': search_query
    }
    return render(request, 'promo/liste_comptables.html', context)


from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from decimal import Decimal

from .models import (
    Paiement, Depense, User, AnneeScolaire,
    student, Classes, Taux, AutreFrais
)


@login_required
def details_comptable(request, id):
    if request.user.role != 'promo':
        return HttpResponseForbidden("Accès interdit")

    comptable = get_object_or_404(
        User,
        id=id,
        role='comptable',
        ecole=request.user.ecole
    )

    annee_id = request.session.get('annee_scolaire')
    annee = get_object_or_404(
        AnneeScolaire,
        id=annee_id,
        ecole=request.user.ecole
    )

    # 🔍 Filtres GET
    filtre_mois = request.GET.get('mois')
    filtre_classe = request.GET.get('classe')
    filtre_objet = request.GET.get('object_paiement')
    filtre_recu = request.GET.get('recu')
    filtre_motif = request.GET.get('motif')

    filtre_mois_dep = request.GET.get('mois_dep')
    filtre_description = request.GET.get('description')

    # 📦 Paiements filtrés
    paiements = None
    if filtre_mois or filtre_classe or filtre_objet or filtre_recu or filtre_motif:
        paiements = Paiement.objects.filter(
            created_by=comptable,
            annee_scolaire=annee
        )
        if filtre_mois:
            paiements = paiements.filter(mois__iexact=filtre_mois)
        if filtre_classe:
            paiements = paiements.filter(classe_id=filtre_classe)
        if filtre_objet:
            paiements = paiements.filter(object_paiement__icontains=filtre_objet)
        if filtre_recu in ['True', 'False']:
            paiements = paiements.filter(recu=(filtre_recu == 'True'))
        if filtre_motif:
            paiements = paiements.filter(object_paiement=filtre_motif)

    # 📦 Dépenses filtrées
    depenses = None
    if filtre_mois_dep or filtre_description:
        depenses = Depense.objects.filter(
            created_by=comptable,
            annee_scolaire=annee
        )
        if filtre_mois_dep and filtre_mois_dep.isdigit():
            depenses = depenses.filter(date_depense__month=int(filtre_mois_dep))
        if filtre_description:
            depenses = depenses.filter(description__description__icontains=filtre_description)

    # 📚 Données générales
    eleves = student.objects.filter(created_by=comptable, annee_scolaire=annee)
    classes = Classes.objects.filter(created_by=comptable, annee_scolaire=annee)
    taux = Taux.objects.filter(created_by=comptable, annee_scolaire=annee)
    autres_frais = AutreFrais.objects.filter(created_by=comptable, annee_scolaire=annee)

    # 📊 Statistiques globales
    total_eleves = eleves.count()
    total_classes = classes.count()

    paiements_globaux = Paiement.objects.filter(created_by=comptable, annee_scolaire=annee, recu=True)
    depenses_globales = Depense.objects.filter(created_by=comptable, annee_scolaire=annee)

    total_cdf = paiements_globaux.aggregate(total=Sum('montant_cdf_brut'))['total'] or Decimal('0')
    total_usd = paiements_globaux.aggregate(total=Sum('montant_usd_brut'))['total'] or Decimal('0')
    total_depenses = depenses_globales.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    total_depense_usd = depenses_globales.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0')

    solde_net = total_cdf - total_depenses
    solde_net_usd = total_usd - total_depense_usd

    paiements_par_mois = paiements_globaux.filter(devise='cdf').values('mois').annotate(
        total=Sum('montant_cdf_brut')
    ).order_by('mois')

    # 📅 Liste des mois (noms)
    MOIS_NOMS = [
        "janvier", "février", "mars", "avril", "mai", "juin",
        "juillet", "août", "septembre", "octobre", "novembre", "décembre"
    ]

    # 🧠 Motifs dynamiques selon mois sélectionné
    motifs_queryset = Paiement.objects.filter(
        created_by=comptable,
        annee_scolaire=annee
    )
    if filtre_mois:
        motifs_queryset = motifs_queryset.filter(mois__iexact=filtre_mois)

    motifs_existants = motifs_queryset.exclude(object_paiement__isnull=True).exclude(object_paiement__exact="").values_list('object_paiement', flat=True).distinct()

    # 🧠 Descriptions dynamiques selon mois sélectionné
    descriptions_queryset = Depense.objects.filter(
        created_by=comptable,
        annee_scolaire=annee
    )
    if filtre_mois_dep and filtre_mois_dep.isdigit():
        descriptions_queryset = descriptions_queryset.filter(date_depense__month=int(filtre_mois_dep))

    descriptions_existantes = descriptions_queryset.exclude(description__description__isnull=True).exclude(description__description__exact="").values_list('description__description', flat=True).distinct()

    # 📦 Contexte
    context = {
        'comptable': comptable,
        'paiements': paiements,
        'depenses': depenses,
        'eleves': eleves,
        'classes': classes,
        'taux': taux,
        'autres_frais': autres_frais,
        'total_eleves': total_eleves,
        'total_classes': total_classes,
        'total_cdf': total_cdf,
        'total_usd': total_usd,
        'total_depenses': total_depenses,
        'total_depense_usd': total_depense_usd,
        'solde_net': solde_net,
        'solde_net_usd': solde_net_usd,
        'paiements_par_mois': paiements_par_mois,
        'filtre_mois': filtre_mois,
        'filtre_classe': filtre_classe,
        'filtre_objet': filtre_objet,
        'filtre_recu': filtre_recu,
        'filtre_motif': filtre_motif,
        'filtre_mois_dep': filtre_mois_dep,
        'filtre_description': filtre_description,
        'mois_noms': MOIS_NOMS,
        'motifs_existants': motifs_existants,
        'descriptions_existantes': descriptions_existantes,
    }

    return render(request, 'promo/details_comptable.html', context)



from academy.models import Paiement, Depense, AnneeScolaire, User


@login_required
def dashboard_promo(request):
    user = request.user

    if user.role != 'promo':
        return HttpResponseForbidden("Accès refusé")

    # 🎯 Récupération des années disponibles
    annees = AnneeScolaire.objects.filter(
        ecole=user.ecole,
        active=True
    ).order_by('nom_annee')

    # 📌 Sélection d’une année scolaire
    annee_selectionnee = None
    annee_id = request.session.get('annee_scolaire')

    if annee_id:
        annee_selectionnee = annees.filter(id=annee_id).first()

    if not annee_selectionnee and annees.exists():
        annee_selectionnee = annees.first()
        request.session['annee_scolaire'] = annee_selectionnee.id
        messages.info(request, f"Année scolaire par défaut sélectionnée : {annee_selectionnee.nom_annee}")

    if not annee_selectionnee:
        messages.error(request, "Aucune année scolaire disponible pour votre école.")
        return redirect('login')

    # 👥 Comptables rattachés à cette école
    comptables = User.objects.filter(role='comptable', ecole=user.ecole)

    # 💰 Paiements et dépenses de ces comptables
    paiements = Paiement.objects.filter(
        annee_scolaire=annee_selectionnee,
        created_by__in=comptables,
        recu=True
    )

    depenses = Depense.objects.filter(
        annee_scolaire=annee_selectionnee,
        created_by__in=comptables
    )

    # 💵 Totaux globaux (sans filtre de devise)
    total_cdf = paiements.aggregate(total=Sum('montant_cdf_brut'))['total'] or Decimal('0')
    total_usd = paiements.aggregate(total=Sum('montant_usd_brut'))['total'] or Decimal('0')

    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or Decimal('0')
    solde_net = total_cdf - total_depenses

    total_depense_usd = depenses.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0')
    solde_net_usd = total_usd - total_depense_usd

    # 📊 Répartition mensuelle
    paiements_par_mois = paiements.filter(devise='cdf').values('mois').annotate(
        total=Sum('montant_cdf_brut')
    ).order_by('mois')

    # 📈 Alertes de gestion
    alertes = []
    if total_depenses > total_cdf:
        alertes.append("❗ Dépenses CDF supérieures aux revenus.")
    elif total_depenses < Decimal('0.7') * total_cdf:
        alertes.append("✅ Dépenses CDF bien maîtrisées.")

    # 🧍 Répartition par comptable
    repartition_comptable = []
    for comptable in comptables:
        total_cdf_comptable = paiements.filter(
            created_by=comptable
        ).aggregate(total=Sum('montant_cdf_brut'))['total'] or Decimal('0')

        total_usd_comptable = paiements.filter(
            created_by=comptable
        ).aggregate(total=Sum('montant_usd_brut'))['total'] or Decimal('0')

        repartition_comptable.append({
            'nom': comptable.nom_complet if comptable.nom_complet else comptable.username,
            'total_cdf': total_cdf_comptable,
            'total_usd': total_usd_comptable
        })

    context = {
        'annees': annees,
        'annee': annee_selectionnee,
        'total_cdf': total_cdf,
        'total_usd': total_usd,
        'total_depenses': total_depenses,
        'solde_net': solde_net,
        'solde_net_usd': solde_net_usd,
        'total_depense_usd': total_depense_usd,
        'paiements_par_mois': paiements_par_mois,
        'comptables': comptables,
        'alertes': alertes,
        'repartition_comptable': repartition_comptable,
    }

    return render(request, 'promo/dashboard_promo.html', context)

#admin 
from dal import autocomplete
from .models import User, AnneeScolaire

class UserAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = User.objects.filter(role='comptable')
        ecole_id = self.forwarded.get('ecole', None)
        if ecole_id:
            qs = qs.filter(ecole_id=ecole_id)
        return qs

class AnneeAutocomplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = AnneeScolaire.objects.all()
        ecole_id = self.forwarded.get('ecole', None)
        if ecole_id:
            qs = qs.filter(ecole_id=ecole_id)
        return qs
   
    
from .models import Ecole

@login_required
def index_promo(request):
    if request.user.role != 'promo':
        return HttpResponseForbidden("Accès interdit")

    return render(request, 'promo/index_promo.html', {'user': request.user})

